import pandas as pd
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QTextEdit, QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import Counter

class CompareExcelTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.layout = QVBoxLayout(self)

        self.log = QTextEdit()
        self.log.setReadOnly(True)

        self.btn_compare = QPushButton("شروع مقایسه و رنگی‌سازی تفاوت‌ها")

        self.layout.addWidget(self.btn_compare)
        self.layout.addWidget(self.log)

        self.btn_compare.clicked.connect(self.compare_with_color_coding)

    def log_message(self, message: str):
        self.log.append(message)
        print(message)

    def select_file(self, prompt="یک فایل را انتخاب کنید"):
        file_path, _ = QFileDialog.getOpenFileName(self, prompt, filter="Excel Files (*.xlsx *.xls)")
        return file_path

    def save_file(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "ذخیره فایل خروجی", filter="Excel Files (*.xlsx)")
        return file_path

    def normalize_row(self, row):
        return tuple(str(x).strip() if pd.notna(x) else "" for x in row)

    def compare_with_color_coding(self):
        try:
            file1 = self.select_file("انتخاب فایل اول")
            if not file1:
                self.log_message("هیچ فایلی برای فایل اول انتخاب نشد.")
                return

            file2 = self.select_file("انتخاب فایل دوم")
            if not file2:
                self.log_message("هیچ فایلی برای فایل دوم انتخاب نشد.")
                return

            df1 = pd.read_excel(file1)
            df2 = pd.read_excel(file2)

            if set(df1.columns) != set(df2.columns):
                QMessageBox.critical(self, "خطا", "ستون‌های دو فایل با هم برابر نیستند.")
                return

            df1 = df1[sorted(df1.columns)]
            df2 = df2[sorted(df2.columns)]

            rows1 = df1.apply(self.normalize_row, axis=1).tolist()
            rows2 = df2.apply(self.normalize_row, axis=1).tolist()

            count1 = Counter(rows1)
            count2 = Counter(rows2)

            all_keys = set(count1.keys()) | set(count2.keys())
            results = []

            for row_key in all_keys:
                c1 = count1.get(row_key, 0)
                c2 = count2.get(row_key, 0)
                row_dict = dict(zip(df1.columns, row_key))

                if c1 > 0 and c2 == 0:
                    row_dict["Source"] = "🟥 فقط در فایل اول"
                elif c2 > 0 and c1 == 0:
                    row_dict["Source"] = "🟩 فقط در فایل دوم"
                elif c1 != c2:
                    row_dict["Source"] = "🟡 تعداد تکرار متفاوت"
                else:
                    continue

                results.append(row_dict)

            if not results:
                QMessageBox.information(self, "نتیجه", "هیچ تفاوتی یافت نشد.")
                return

            result_df = pd.DataFrame(results)
            output_path = self.save_file()
            if output_path:
                result_df.to_excel(output_path, index=False)

                wb = load_workbook(output_path)
                ws = wb.active

                color_map = {
                    "🟥 فقط در فایل اول": "FFC7CE",
                    "🟩 فقط در فایل دوم": "C6EFCE",
                    "🟡 تعداد تکرار متفاوت": "FFF2CC",
                }

                for i, row in enumerate(result_df.itertuples(), start=2):
                    label = getattr(row, "Source")
                    fill_color = color_map.get(label, None)
                    if fill_color:
                        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        for col in range(1, len(result_df.columns) + 1):
                            ws.cell(row=i, column=col).fill = fill

                wb.save(output_path)
                QMessageBox.information(self, "موفقیت", f"فایل خروجی با رنگ‌بندی ذخیره شد:\n{output_path}")

        except Exception as e:
            QMessageBox.critical(self, "خطا", str(e))
