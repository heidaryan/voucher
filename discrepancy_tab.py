import pandas as pd
import os
import sys
from PyQt5.QtWidgets import QApplication, QFileDialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import BUILTIN_FORMATS
import shutil
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QPushButton, QTextEdit, QFileDialog, QMessageBox
import os

class ExcelFolderReader:
    def __init__(self, output_folder, target_columns , target_columns_code):

        self.VoucherRow_TS = None
        self.AccountCoding_TS_Level1 = None
        self.AccountCoding_TS_Level2 = None
        self.AccountCoding_TS_Level3 = None
        self.output_folder = output_folder
        self.target_columns = target_columns
        self.target_columns_code = target_columns_code

    
    def open_folder_dialog(self):
        app = QApplication.instance()
        if not app:
            app = QApplication(sys.argv)
            folder_path = QFileDialog.getExistingDirectory(None, " انتخاب فولدر حاوی فایل‌های ریزاسناد و کدینگ (Folder)   ")
            return folder_path



    def read_excel_files(self, folder_path, columns_to_keep_leading_zeros=None, columns_to_keep_float=None):

        if columns_to_keep_leading_zeros is None:
            columns_to_keep_leading_zeros = []
        if columns_to_keep_float is None:
            columns_to_keep_float = []

        # نگاشت نام فایل‌ها به نام متغیرها
        file_map = {
            "VoucherRow_TS.xlsx": "VoucherRow_TS",
            "AccountCoding_TS_Level1.xlsx": "AccountCoding_TS_Level1",
            "AccountCoding_TS_Level2.xlsx": "AccountCoding_TS_Level2",
            "AccountCoding_TS_Level3.xlsx": "AccountCoding_TS_Level3"
        }

        for file_name in os.listdir(folder_path):
            if file_name in file_map:
                file_path = os.path.join(folder_path, file_name)
                try:

                    dtype_dict = {col: str for col in columns_to_keep_leading_zeros}


                    dtype_dict.update({col: float for col in columns_to_keep_float})


                    df = pd.read_excel(file_path, dtype=dtype_dict)


                    setattr(self, file_map[file_name], df)
                    print(f"فایل '{file_name}' با موفقیت در متغیر '{file_map[file_name]}' ذخیره شد.")
                except Exception as e:
                    print(f"خطا در خواندن فایل {file_name}: {e}")

    def display_dataframes(self):

        if self.VoucherRow_TS is not None:
            print("دیتافریم: VoucherRow_TS")
            print(self.VoucherRow_TS.head())
        if self.AccountCoding_TS_Level1 is not None:
            print("دیتافریم: AccountCoding_TS_Level1")
            print(self.AccountCoding_TS_Level1.head())
        if self.AccountCoding_TS_Level2 is not None:
            print("دیتافریم: AccountCoding_TS_Level2")
            print(self.AccountCoding_TS_Level2.head())
        if self.AccountCoding_TS_Level3 is not None:
            print("دیتافریم: AccountCoding_TS_Level3")
            print(self.AccountCoding_TS_Level3.head())

    def clear_output_folder(self):

        if os.path.exists(self.output_folder):
            # حذف تمام محتویات پوشه
            for filename in os.listdir(self.output_folder):
                file_path = os.path.join(self.output_folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)  
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path) 
                except Exception as e:
                    print(f"خطا در حذف فایل {file_path}: {e}")
            print(f"پوشه '{self.output_folder}' خالی شد.")
        else:
            print(f"پوشه '{self.output_folder}' وجود ندارد.")

    def get_rows_with_missing_values(self, df):


        missing_columns = [col for col in self.target_columns if col not in df.columns]
        if missing_columns:
            print(f"ستون‌های زیر در فایل یافت نشدند: {missing_columns}")
            return pd.DataFrame() 


        rows_with_missing = df[df[self.target_columns].isnull().any(axis=1)]
        return rows_with_missing



    def save_missing_rows_with_formatting(self, file_name, missing_rows):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not missing_rows.empty:
            output_path = os.path.join(self.output_folder, f"5_missing_rows_{file_name}.xlsx")
            missing_rows.to_excel(output_path, index=False)
            print(f"ردیف‌های دارای مقادیر خالی برای فایل '{file_name}' ذخیره شد.")
            
            wb = load_workbook(output_path)
            ws = wb.active
            red_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is None:
                        cell.fill = red_fill
            
            wb.save(output_path)
            print(f"قالب‌بندی (رنگ بنفش) برای سلول‌های خالی در فایل '{file_name}' اعمال شد.")
        else:
            print(f"هیچ ردیفی با مقدار خالی در فایل '{file_name}' وجود ندارد.")



    def get_rows_with_missing_values_code(self, df):


        missing_columns = [col for col in self.target_columns_code if col not in df.columns]
        if missing_columns:
            print(f"ستون‌های زیر در فایل یافت نشدند: {missing_columns}")
            return pd.DataFrame() 


        rows_with_missing_code = df[df[self.target_columns_code].isnull().any(axis=1)]
        return rows_with_missing_code


    def save_missing_rows_with_formatting_code(self, file_name, missing_rows_code):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not missing_rows_code.empty:
            output_path = os.path.join(self.output_folder, f"5_missing_rows_{file_name}.xlsx")
            missing_rows_code.to_excel(output_path, index=False)
            print(f"ردیف‌های دارای مقادیر خالی برای فایل '{file_name}' ذخیره شد.")
            
            wb = load_workbook(output_path)
            ws = wb.active
            red_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is None:
                        cell.fill = red_fill
            
            wb.save(output_path)
            print(f"قالب‌بندی (رنگ بنفش) برای سلول‌های خالی در فایل '{file_name}' اعمال شد.")
        else:
            print(f"هیچ ردیفی با مقدار خالی در فایل '{file_name}' وجود ندارد.")



    def check_debtor_credit_balance_html(self, df, html_file_path):

        if 'DebtorAmount' in df.columns and 'CreditorAmount' in df.columns:
            total_debtor = df['DebtorAmount'].sum()
            total_creditor = df['CreditorAmount'].sum()

            if total_debtor == total_creditor:
                message = f"<p style='color:green;'>✅ جمع بدهکار ({total_debtor}) و جمع بستانکار ({total_creditor}) برابر هستند. 😊</p>"
            else:
                message = f"<p style='color:red;'>❌ مغایرت در بدهکار و بستانکار وجود دارد: جمع بدهکار ({total_debtor})، جمع بستانکار ({total_creditor}). 😞</p>"
        else:
            message = "<p style='color:purple;'>ستون‌های 'DebtorAmount' و 'CreditorAmount' در فایل موجود نیستند.</p>"

        # ذخیره پیام در فایل HTML
        with open(html_file_path, 'w', encoding='utf-8') as html_file:
            html_file.write("<html><body>")
            html_file.write(message)
            html_file.write("</body></html>")
        print("پیام بررسی بالانس در فایل HTML ذخیره شد.")



    def check_balance_VoucherNumber(self, df):

        grouped = df.groupby('VoucherNumber').agg({'DebtorAmount': 'sum', 'CreditorAmount': 'sum'}).reset_index()
        grouped['تراز'] = grouped['DebtorAmount'] - grouped['CreditorAmount']
        grouped['وضعیت'] = grouped['تراز'].apply(lambda x: 'بالانس' if x == 0 else 'نامتعادل')
        unbalanced_rows_vo = grouped[grouped['وضعیت'] == 'نامتعادل']

        if unbalanced_rows_vo.empty:
            print("سطر نامتعادل وجود ندارد.")
        else:
            print("سطرهای نامتعادل:")
            print(unbalanced_rows_vo)

        return unbalanced_rows_vo


    def save_balance_VoucherNumber_row(self, file_name, unbalanced_rows_vo):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not unbalanced_rows_vo.empty:
            unbalanced_rows_vo['CreditorAmount']= unbalanced_rows_vo['CreditorAmount'].apply(lambda x: '{:.0f}'.format(x))
            unbalanced_rows_vo['DebtorAmount']= unbalanced_rows_vo['DebtorAmount'].apply(lambda x: '{:.0f}'.format(x))
            unbalanced_rows_vo['تراز']= unbalanced_rows_vo['تراز'].apply(lambda x: '{:.0f}'.format(x))
            output_path = os.path.join(self.output_folder, f"8_unbalanced_vouchor_{file_name}.xlsx")
            unbalanced_rows_vo.to_excel(output_path, index=False)
            print(f"سطرهای نامتعادل برای فایل '{file_name}' ذخیره شد.")
        else:
            print(f"هیچ سطر نامتعادلی در فایل '{file_name}' وجود ندارد.")



    def check_VoucherType_Flag_consistency(self, df):
        if 'VoucherType_Flag' not in df.columns or 'VoucherNumber' not in df.columns:
            print("ستون‌های 'VoucherType_Flag' یا 'VoucherNumber' در دیتافریم موجود نیستند.")
            return pd.DataFrame()
        inconsistent_records_num = df.groupby('VoucherNumber').filter(
            lambda x: x['VoucherType_Flag'].nunique() > 1
        )

        if inconsistent_records_num.empty:
            print("✅ همه اسناد تنها یک تاریخ دارند.")
        else:
            print("❌ اسنادی با بیش از یک تاریخ یافت شدند:")
            print(inconsistent_records_num)

        return inconsistent_records_num

    def save_inconsistent_num_flag(self, file_name, inconsistent_records_num):
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not inconsistent_records_num.empty:
            output_path = os.path.join(self.output_folder, f"unmach_vouch_flag_{file_name}.xlsx")
            inconsistent_records_num.to_excel(output_path, index=False)
            print(f"رکوردهای دارای flag ناسازگار در فایل '{output_path}' ذخیره شدند.")
        else:
            print("همه اسناد تنها یک flag داشتند. نیازی به ذخیره فایل نیست.")

 

    def validate_description_rows(self, df):
        if 'VoucherType_Flag' not in df.columns or 'DescriptionRow' not in df.columns:
            print("ستون‌های 'VoucherType_Flag' یا 'DescriptionRow' در دیتافریم موجود نیستند.")
            return pd.DataFrame()  

        rows_with_flag_2 = df[df['VoucherType_Flag'] == 2]

        valid_keywords = ['افتتاحیه', 'افتتاح', 'افتاح', 'افتاحیه']

        invalid_rows = rows_with_flag_2[~rows_with_flag_2['DescriptionRow'].str.contains('|'.join(valid_keywords), na=False)]

        return invalid_rows

    def save_invalid_description_rows(self, file_name, invalid_rows):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not invalid_rows.empty:
            output_path = os.path.join(self.output_folder, f"4_invalid_description_rows_{file_name}.xlsx")
            invalid_rows.to_excel(output_path, index=False)
            print(f"رکوردهای دارای DescriptionRow مغایر برای فایل '{file_name}' ذخیره شدند.")
        else:
            print(f"هیچ رکورد مغایری در ستون 'DescriptionRow' برای فایل '{file_name}' وجود ندارد.")



    def check_one_to_one_relationship(self, df, column1, column2, file_name):
        if column1 not in df.columns or column2 not in df.columns:
            print(f"ستون‌های '{column1}' یا '{column2}' در دیتافریم موجود نیستند.")
            return pd.DataFrame()

        grouped1 = df.groupby(column1)[column2].nunique().reset_index()
        non_unique_codes1 = grouped1[grouped1[column2] > 1]

        grouped2 = df.groupby(column2)[column1].nunique().reset_index()
        non_unique_codes2 = grouped2[grouped2[column1] > 1]

        result = []
        
        for code in non_unique_codes1[column1]:
            names = df[df[column1] == code][column2].unique()
            for name in names:
                result.append([code, name])
        for code in non_unique_codes2[column2]:
            names = df[df[column2] == code][column1].unique()
            for name in names:
                result.append([name, code])

        result_df = pd.DataFrame(result, columns=[column1, column2])

        if not result_df.empty:
            self.save_one_to_one_results(file_name, result_df)
        else:
            print(f"هیچ مغایرتی در رابطه یک به یک بین '{column1}' و '{column2}' یافت نشد.")
        
        return result_df


    def save_one_to_one_results(self, file_name, result_df):
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not result_df.empty:
            output_path = os.path.join(self.output_folder, f"6_one_to_one_relationship_{file_name}.xlsx")
            result_df.to_excel(output_path, index=False)
            print(f"نتایج در فایل '{output_path}' ذخیره شد.")


    def check_date_consistency(self, df):

        if 'PersianVoucherDate' not in df.columns or 'VoucherNumber' not in df.columns:
            print("ستون‌های 'PersianVoucherDate' یا 'VoucherNumber' در دیتافریم موجود نیستند.")
            return pd.DataFrame()

        inconsistent_records = df.groupby('VoucherNumber').filter(
            lambda x: x['PersianVoucherDate'].nunique() > 1
        )

        if inconsistent_records.empty:
            print("✅ همه اسناد تنها یک تاریخ دارند.")
        else:
            print("❌ اسنادی با بیش از یک تاریخ یافت شدند:")
            print(inconsistent_records)

        return inconsistent_records

    def save_inconsistent_dates(self, file_name, inconsistent_records):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not inconsistent_records.empty:
            output_path = os.path.join(self.output_folder, f"9_unmach_dates_{file_name}.xlsx")
            inconsistent_records.to_excel(output_path, index=False)
            print(f"رکوردهای دارای تاریخ‌های ناسازگار در فایل '{output_path}' ذخیره شدند.")
        else:
            print("همه اسناد تنها یک تاریخ داشتند. نیازی به ذخیره فایل نیست.")


    def check_VoucherType_Flag_consistency(self, df):
        if 'VoucherType_Flag' not in df.columns or 'VoucherNumber' not in df.columns:
            print("ستون‌های 'VoucherType_Flag' یا 'VoucherNumber' در دیتافریم موجود نیستند.")
            return pd.DataFrame()
        inconsistent_records_num = df.groupby('VoucherNumber').filter(
            lambda x: x['VoucherType_Flag'].nunique() > 1
        )

        if inconsistent_records_num.empty:
            print("✅ همه اسناد تنها یک تاریخ دارند.")
        else:
            print("❌ اسنادی با بیش از یک تاریخ یافت شدند:")
            print(inconsistent_records_num)

        return inconsistent_records_num


    def save_inconsistent_num_flag(self, file_name, inconsistent_records_num):
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not inconsistent_records_num.empty:
            output_path = os.path.join(self.output_folder, f"10_unmach_vouch_flag_{file_name}.xlsx")
            inconsistent_records_num.to_excel(output_path, index=False)
            print(f"رکوردهای دارای flag ناسازگار در فایل '{output_path}' ذخیره شدند.")
        else:
            print("همه اسناد تنها یک flag داشتند. نیازی به ذخیره فایل نیست.")


    def get_rows_with_invalid_code_length(self, df, column_name):
        if column_name not in df.columns:
            print(f"ستون '{column_name}' در فایل موجود نیست.")
            return pd.DataFrame()  

        df['Code_Length'] = df[column_name].astype(str).str.len()

        common_length = df['Code_Length'].mode()[0]

        invalid_rows = df[df['Code_Length'] != common_length]
        df.drop(columns=['Code_Length'], inplace=True)

        return invalid_rows


    def save_invalid_rows(self, file_name, invalid_rows, column_name):
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not invalid_rows.empty:
            output_path = os.path.join(self.output_folder, f"{column_name}_invalid_code_length_{file_name}.xlsx")
            invalid_rows.to_excel(output_path, index=False)
            print(f"ردیف‌های با طول نامعتبر در ستون '{column_name}' برای فایل '{file_name}' ذخیره شد.")
        else:
            print(f"همه ردیف‌ها در فایل '{file_name}' طول معتبر دارند.")



    def check_persian_date_format(self, df, column_name):

        if column_name not in df.columns:
            print(f"ستون '{column_name}' در دیتافریم موجود نیست.")
            return pd.DataFrame() 

        persian_date_pattern = r'^[0-9]{4}/[0-9]{2}/[0-9]{2}$'
        
        invalid_dates = df[~df[column_name].astype(str).str.match(persian_date_pattern)]
        
        return invalid_dates

    def save_invalid_dates(self, file_name, invalid_dates):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not invalid_dates.empty:
            output_path = os.path.join(self.output_folder, f"3_invalid_dates_{file_name}.xlsx")
            invalid_dates.to_excel(output_path, index=False)
            print(f"تاریخ‌های غیرمعتبر برای فایل '{file_name}' ذخیره شد.")
        else:
            print(f"همه تاریخ‌ها در ستون '{file_name}' فرمت فارسی دارند.")



    def check_balance_flag(self, df):

        grouped = df.groupby('VoucherType_Flag').agg({'DebtorAmount': 'sum', 'CreditorAmount': 'sum'}).reset_index()
        grouped['تراز'] = grouped['DebtorAmount'] - grouped['CreditorAmount']
        grouped['وضعیت'] = grouped['تراز'].apply(lambda x: 'بالانس' if x == 0 else 'نامتعادل')
        unbalanced_rows = grouped[grouped['وضعیت'] == 'نامتعادل']

        if unbalanced_rows.empty:
            print("سطر نامتعادل وجود ندارد.")
        else:
            print("سطرهای نامتعادل:")
            print(unbalanced_rows)

        return unbalanced_rows

    def save_balance_flag_rows(self, file_name, unbalanced_rows):
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not unbalanced_rows.empty:
            unbalanced_rows['CreditorAmount']= unbalanced_rows['CreditorAmount'].apply(lambda x: '{:.0f}'.format(x))
            unbalanced_rows['DebtorAmount']= unbalanced_rows['DebtorAmount'].apply(lambda x: '{:.0f}'.format(x))
            unbalanced_rows['تراز']= unbalanced_rows['تراز'].apply(lambda x: '{:.0f}'.format(x))

            output_path = os.path.join(self.output_folder, f"7_unbalanced_flag_rows_{file_name}.xlsx")
            unbalanced_rows.to_excel(output_path, index=False ,  float_format='%.0f')
            print(f"سطرهای نامتعادل برای فایل '{file_name}' ذخیره شد.")
        else:
            print(f"هیچ سطر نامتعادلی در فایل '{file_name}' وجود ندارد.")




    def check_code_discrepancies(self, voucher_df, coding_df, output_filename):

        voucher_df['Code'] = voucher_df['Code'].astype(str).str.strip()
        coding_df['Code'] = coding_df['Code'].astype(str).str.strip()


        unique_voucher_codes = voucher_df['Code'].drop_duplicates()


        missing_in_coding = unique_voucher_codes[~unique_voucher_codes.isin(coding_df['Code'])]


        missing_in_voucher = coding_df[~coding_df['Code'].isin(unique_voucher_codes)]


        discrepancies = pd.DataFrame()

        if not missing_in_coding.empty:
            discrepancies = pd.concat([discrepancies, pd.DataFrame({'Code (Missing in Coding)': missing_in_coding})], ignore_index=True)

        if not missing_in_voucher.empty:
            discrepancies = pd.concat([discrepancies, pd.DataFrame({'Code (Unused in Voucher)': missing_in_voucher['Code']})], ignore_index=True)
        self.save_code_discrepancies(discrepancies, output_filename)

    def save_code_discrepancies(self, discrepancies_df, output_filename):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not discrepancies_df.empty:
            output_path = os.path.join(self.output_folder, output_filename)
            discrepancies_df.to_excel(output_path, index=False)
            print(f"مغایرت‌ها در فایل '{output_path}' ذخیره شدند.")
        else:
            print("هیچ مغایرتی یافت نشد. نیازی به ذخیره فایل نیست.")



    def check_code_parent_child_dynamic(self, parent_df, child_df, output_filename):
        discrepancies = self.find_code_discrepancies(parent_df, child_df)


        self.save_discrepancies_to_excel(discrepancies, output_filename)

    def find_code_discrepancies(self, parent_df, child_df):

        parent_df['Code'] = parent_df['Code'].astype(str).fillna("")
        child_df['Code'] = child_df['Code'].astype(str).fillna("")

        discrepancies = child_df[~child_df['Code'].str[:len(parent_df['Code'][0])].isin(parent_df['Code'])]
        return discrepancies

    def save_discrepancies_to_excel(self, discrepancies, output_filename):
        if not discrepancies.empty:
            output_path = os.path.join(self.output_folder, output_filename)
            discrepancies.to_excel(output_path, index=False)
            print(f"مغایرت‌ها در فایل {output_filename} ذخیره شدند.")
        else:
            print(f"هیچ مغایرتی یافت نشد.")





    def get_rows_with_invalid_code_lengthmat(self, df, column_name):
        if column_name not in df.columns:
            print(f"ستون '{column_name}' در فایل موجود نیست.")
            return pd.DataFrame()

        # فیلتر کردن مقادیر غیر null
        filtered_df = df[df[column_name].notna()]

        # محاسبه طول کدها
        filtered_df['Code_Length'] = filtered_df[column_name].astype(str).apply(len)

        # محاسبه مود
        modes = filtered_df['Code_Length'].mode()
        if not modes.empty:
            common_length = modes[0]
            # یافتن سطرهای با طول کد نامعتبر
            invalid_rows_ = filtered_df[filtered_df['Code_Length'] != common_length]
        else:
            print("هیچ مقداری برای محاسبه مود یافت نشد.")
            filtered_df.drop(columns=['Code_Length'], inplace=True)
            return pd.DataFrame()

        # حذف ستون کمکی
        filtered_df.drop(columns=['Code_Length'], inplace=True)

        return invalid_rows_


    def save_invalid_rows_(self, file_name, invalid_rows_, column_name):

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        if not invalid_rows_.empty:
            output_path = os.path.join(self.output_folder, f"{column_name}_invalid_code_length_{file_name}.xlsx")
            invalid_rows_.to_excel(output_path, index=False)
            print(f"ردیف‌های با طول نامعتبر در ستون '{column_name}' برای فایل '{file_name}' ذخیره شد.")
        else:
            print(f"همه ردیف‌ها در فایل '{file_name}' طول معتبر دارند.")

 
 
    def is_matrix_project(self, df, threshold=0.1):

        non_empty_columns = (df.notna().sum() / len(df)) > threshold
        return non_empty_columns.any()


class DiscrepancyTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        # مسیر دقیق پوشه خروجی گزارش‌ها
        self.output_folder = r"D:/vochor/vsTokenoutput_errors"
        
        self.target_columns = [
            'Code', 'Name', 'DebtorAmount', 'CreditorAmount', 'VoucherNumber',
            'RowNumber', 'PersianVoucherDate', 'VoucherType_Flag'
        ]
        self.target_columns_code = ['Code', 'Name']

        # نمونه ExcelFolderReader با مسیر خروجی درست
        self.reader = ExcelFolderReader(self.output_folder, self.target_columns, self.target_columns_code)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.label_status = QLabel("وضعیت: آماده")
        self.btn_select_folder = QPushButton("انتخاب فولدر فایل‌های اکسل")
        self.btn_run_check = QPushButton("اجرای بررسی مغایرت‌ها")
        self.text_log = QTextEdit()
        self.text_log.setReadOnly(True)

        layout.addWidget(self.label_status)
        layout.addWidget(self.btn_select_folder)
        layout.addWidget(self.btn_run_check)
        layout.addWidget(self.text_log)

        self.setLayout(layout)

        self.btn_select_folder.clicked.connect(self.select_folder)
        self.btn_run_check.clicked.connect(self.run_checks)

        self.selected_folder = None

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "انتخاب فولدر حاوی فایل‌های اکسل")
        if folder:
            self.selected_folder = folder
            self.label_status.setText(f"فولدر انتخاب شد:\n{folder}")
            self.log(f"فولدر انتخاب شده: {folder}")
        else:
            self.log("هیچ فولدری انتخاب نشد.")

    def log(self, message):
        self.text_log.append(message)

    def run_checks(self):
        if not self.selected_folder:
            QMessageBox.warning(self, "هشدار", "لطفاً ابتدا فولدر حاوی فایل‌های اکسل را انتخاب کنید.")
            return

        self.log("شروع خواندن فایل‌ها...")

        columns_to_keep_leading_zeros = ['Code', 'Matrix_1_Code', 'Matrix_2_Code']
        columns_to_keep_float = ['DebtorAmount', 'CreditorAmount']

        # خواندن فایل‌های اکسل
        self.reader.read_excel_files(self.selected_folder, columns_to_keep_leading_zeros, columns_to_keep_float)
        self.log("خواندن فایل‌ها انجام شد.")

        if self.reader.VoucherRow_TS is not None:
            self.log("فایل VoucherRow_TS خوانده شد.")
            self.reader.clear_output_folder()

            missing_rows = self.reader.get_rows_with_missing_values(self.reader.VoucherRow_TS)
            self.reader.save_missing_rows_with_formatting("VoucherRow_TS", missing_rows)

            html_file_path = os.path.join(self.output_folder, "1_debtor_credit_balance.html")
            self.reader.check_debtor_credit_balance_html(self.reader.VoucherRow_TS, html_file_path)

            unbalanced_rows_VoucherNumber = self.reader.check_balance_VoucherNumber(self.reader.VoucherRow_TS)
            self.reader.save_balance_VoucherNumber_row("VoucherRow_TS", unbalanced_rows_VoucherNumber)

            unmach_vou_flag_num = self.reader.check_VoucherType_Flag_consistency(self.reader.VoucherRow_TS)
            self.reader.save_inconsistent_num_flag("VoucherRow_TS", unmach_vou_flag_num)

            invalid_description_rows = self.reader.validate_description_rows(self.reader.VoucherRow_TS)
            self.reader.save_invalid_description_rows("VoucherRow_TS", invalid_description_rows)

            inconsistent_records = self.reader.check_date_consistency(self.reader.VoucherRow_TS)
            self.reader.save_inconsistent_dates("VoucherRow_TS", inconsistent_records)

            invalid_rows = self.reader.get_rows_with_invalid_code_length(self.reader.VoucherRow_TS, 'Code')
            self.reader.save_invalid_rows("VoucherRow_TS", invalid_rows, 'Code')

            invalid_dates = self.reader.check_persian_date_format(self.reader.VoucherRow_TS, 'PersianVoucherDate')
            self.reader.save_invalid_dates("VoucherRow_TS", invalid_dates)

            result_df = self.reader.check_one_to_one_relationship(self.reader.VoucherRow_TS, 'Code', 'Name', "VoucherRow_TS")
            self.reader.save_one_to_one_results("VoucherRow_TS", result_df)

            unbalanced_rows_flag = self.reader.check_balance_flag(self.reader.VoucherRow_TS)
            self.reader.save_balance_flag_rows("VoucherRow_TS", unbalanced_rows_flag)

            invalid_rows_1 = self.reader.get_rows_with_invalid_code_length(self.reader.AccountCoding_TS_Level1, 'Code')
            self.reader.save_invalid_rows("AccountCoding_TS_Level1", invalid_rows_1, 'Code')

            if self.reader.is_matrix_project(self.reader.VoucherRow_TS):
                invalid_rows_021 = self.reader.get_rows_with_invalid_code_lengthmat(self.reader.VoucherRow_TS, 'Matrix_1_Code')
                self.reader.save_invalid_rows_("VoucherRow_TS", invalid_rows_021, 'Matrix_1_Code')

                invalid_rows_021 = self.reader.get_rows_with_invalid_code_lengthmat(self.reader.VoucherRow_TS, 'Matrix_2_Code')
                self.reader.save_invalid_rows_("VoucherRow_TS", invalid_rows_021, 'Matrix_2_Code')
            else:
                self.log("دیتافریم ماتریسی نیست، نیازی به بررسی طول کد نیست.")
        else:
            self.log("فایل VoucherRow_TS یافت نشد یا خوانده نشده است.")

        # بررسی AccountCoding_TS_Level1
        if self.reader.AccountCoding_TS_Level1 is not None:
            missing_rows_code1 = self.reader.get_rows_with_missing_values_code(self.reader.AccountCoding_TS_Level1)
            self.reader.save_missing_rows_with_formatting_code("AccountCoding_TS_Level1", missing_rows_code1)

            result_df_1 = self.reader.check_one_to_one_relationship(self.reader.AccountCoding_TS_Level1, 'Code', 'Name', "AccountCoding_TS_Level1")
            self.reader.save_one_to_one_results("AccountCoding_TS_Level1", result_df_1)

            invalid_rows_1 = self.reader.get_rows_with_invalid_code_length(self.reader.AccountCoding_TS_Level1, 'Code')
            self.reader.save_invalid_rows("AccountCoding_TS_Level1", invalid_rows_1, 'Code')
        else:
            self.log("فایل AccountCoding_TS_Level1 یافت نشد یا خوانده نشده است.")

        # بررسی AccountCoding_TS_Level2
        if self.reader.AccountCoding_TS_Level2 is not None:
            missing_rows_code2 = self.reader.get_rows_with_missing_values_code(self.reader.AccountCoding_TS_Level2)
            self.reader.save_missing_rows_with_formatting_code("AccountCoding_TS_Level2", missing_rows_code2)

            result_df_2 = self.reader.check_one_to_one_relationship(self.reader.AccountCoding_TS_Level2, 'Code', 'Name', "AccountCoding_TS_Level2")
            self.reader.save_one_to_one_results("AccountCoding_TS_Level2", result_df_2)

            invalid_rows_2 = self.reader.get_rows_with_invalid_code_length(self.reader.AccountCoding_TS_Level2, 'Code')
            self.reader.save_invalid_rows("AccountCoding_TS_Level2", invalid_rows_2, 'Code')
        else:
            self.log("فایل AccountCoding_TS_Level2 یافت نشد یا خوانده نشده است.")

        # بررسی AccountCoding_TS_Level3
        if self.reader.AccountCoding_TS_Level3 is not None:
            missing_rows_code3 = self.reader.get_rows_with_missing_values_code(self.reader.AccountCoding_TS_Level3)
            self.reader.save_missing_rows_with_formatting_code("AccountCoding_TS_Level3", missing_rows_code3)

            invalid_rows_3 = self.reader.get_rows_with_invalid_code_length(self.reader.AccountCoding_TS_Level3, 'Code')
            self.reader.save_invalid_rows("AccountCoding_TS_Level3", invalid_rows_3, 'Code')
        else:
            self.log("فایل AccountCoding_TS_Level3 یافت نشد یا خوانده نشده است.")

        if self.reader.AccountCoding_TS_Level3 is not None:
            self.log("در حال بررسی AccountCoding_TS_Level3...")
            self.reader.check_code_discrepancies(self.reader.VoucherRow_TS, self.reader.AccountCoding_TS_Level3, "11_Discrepancies_Level3_Report.xlsx")
        elif self.reader.AccountCoding_TS_Level2 is not None:
            self.log("در حال بررسی AccountCoding_TS_Level2...")
            self.reader.check_code_discrepancies(self.reader.VoucherRow_TS, self.reader.AccountCoding_TS_Level2, "11_Discrepancies_Level2_Report.xlsx")
        else:
            self.log("هیچ دیتافریمی برای بررسی موجود نیست.")

        if self.reader.AccountCoding_TS_Level3 is not None:
            self.log("در حال بررسی AccountCoding_TS_Level3...")
            self.reader.check_code_parent_child_dynamic(self.reader.AccountCoding_TS_Level1, self.reader.AccountCoding_TS_Level3, "12Parent_Child_Discrepancies_Level3_Report.xlsx")

        if self.reader.AccountCoding_TS_Level2 is not None:
            self.log("در حال بررسی AccountCoding_TS_Level2...")
            self.reader.check_code_parent_child_dynamic(self.reader.AccountCoding_TS_Level1, self.reader.AccountCoding_TS_Level2, "12Parent_Child_Discrepancies_Level2_Report.xlsx")
        else:
            self.log("هیچ دیتافریمی برای بررسی موجود نیست.")

        self.label_status.setText("وضعیت: بررسی کامل شد.")
        self.log("تمام بررسی‌ها انجام و گزارش‌ها ذخیره شدند.")
